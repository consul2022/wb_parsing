from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from app.core.config import WB_PARSING_BASE_URL, MAX_PAGES
from app.services.parser import start_driver, scroll_page

from openpyxl import Workbook
from io import BytesIO
def search_links(driver):
    cards = driver.find_elements(By.CSS_SELECTOR, "article.product-card a.product-card__link")
    links = []
    for card in cards:
        links.append(card.get_attribute("href"))
        if len(links) >= MAX_PAGES:
            break
    return links


def get_info_page(driver, link):
    driver.get(link)
    scroll_page(driver)
    info = {'url': link}

    article = driver.find_element(By.XPATH,
                                  '//*[@id="reactContainers"]/div[2]/div/div[3]/div[2]/div[3]/div/div/table/tbody/tr[1]/td/button/span')
    info['article'] = article.text

    name = driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[3]/div[2]/div[1]/div/div[1]/h3')
    info['name'] = name.text

    price = driver.find_element(By.XPATH,
                                '//*[@id="reactContainers"]/div[2]/div/div[3]/div[3]/div/div/div[1]/div/div/div/div/div/span[1]/ins')
    info['price'] = price.text

    info['images'] = []
    images = driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[3]/div[1]/div/div/div[1]/div')
    for image in images.find_elements(By.TAG_NAME, 'img'):
        info['images'].append(image.get_attribute("src"))

    seller_name = driver.find_element(By.XPATH,
                                      '//*[@id="reactContainers"]/div[2]/div/div[3]/div[3]/div/div/div[5]/section/div/div/div/a/div[2]/div/div/span[1]')
    info['seller_name'] = seller_name.text

    seller_link = driver.find_element(By.XPATH,
                                      '//*[@id="reactContainers"]/div[2]/div/div[3]/div[3]/div/div/div[5]/section/div/div/div/a')
    info['seller_link'] = seller_link.get_attribute("href")

    info['sizes'] = []
    sizes = driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[3]/div[2]/div[2]/div[2]/ul')
    for size in sizes.find_elements(By.XPATH, './/li//button//span[1]'):
        info['sizes'].append(size.text)

    info['sizes_available'] = []
    sizes_available = driver.find_element(By.XPATH,
                                          '//*[@id="reactContainers"]/div[2]/div/div[3]/div[2]/div[2]/div[2]/ul')
    for size in sizes_available.find_elements(By.XPATH, './/li[contains(@class,"sizeActive")]//button//span[1]'):
        info['sizes_available'].append(size.text)
    if not info["sizes_available"]:
        info['sizes_available'] = info["sizes"]

    rating = driver.find_element(By.XPATH, '//*[@id="product-feedbacks"]/div[2]/div[1]/div[1]/div[1]/b')
    info['rating'] = rating.text

    reviews_amount = driver.find_element(By.XPATH, '//*[@id="product-feedbacks"]/div[2]/div[1]/div[1]/a')
    info['reviews_amount'] = reviews_amount.text.split()[0]

    wait = WebDriverWait(driver, 15)
    btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='reactContainers']/div[2]/div/div[3]/div[2]/div[3]/div/button")
    ))

    btn.click()
    modal = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//div[contains(@class,'mo-modal__paper') and .//h2[contains(.,'Характеристики и описание')]]")
    ))

    description = modal.find_element(By.XPATH, '//*[@id="section-description"]//p')
    info['description'] = description.text.strip()

    tables = modal.find_elements(By.XPATH, ".//section[@data-testid='product_additional_information']//table")
    info['characteristic'] = {}
    for table in tables:
        caption = table.find_element(By.XPATH, './/caption')
        info['characteristic'][caption.text] = []
        rows = table.find_elements(By.XPATH, './/tbody/tr')
        for row in rows:
            key = row.find_element(By.XPATH, './/th').text.strip()
            value = row.find_element(By.XPATH, './/td').text.strip()
            info['characteristic'][caption.text].append({"key" : key, "value" : value})

    return info

def get_search_results(search):
    driver = start_driver()
    driver.get(f"{WB_PARSING_BASE_URL}/catalog/0/search.aspx?search={search}")
    scroll_page(driver)
    links = search_links(driver)
    total_info = []
    for link in links:
        total_info.append(get_info_page(driver, link))
    driver.quit()
    return total_info

def from_dict_get_excel(total_info, min_rating, max_price):
    filtered_info = []
    for info in total_info:
        if min_rating is not None and float(info['rating'].replace(',', '.')) < min_rating:
            continue
        if max_price is not None and float(info['price'].r) > max_price:
            continue
        filtered_info.append(info)

    wb = Workbook()
    ws = wb.active
    ws.title = "WB Products"
    ws['A1'] = 'Ссылка на товар'
    ws['B1'] = 'Артикул'
    ws['C1'] = 'Название'
    ws['D1'] = 'Цена'
    ws['E1'] = 'Ссылки на изображения'
    ws['F1'] = 'Описание'
    ws['G1'] = 'Основная информация'
    ws['H1'] = 'Дополнительная информация'
    ws['I1'] = 'Название селлера'
    ws['J1'] = 'Ссылка на селлера'
    ws['K1'] = 'Размеры товаров'
    ws['L1'] = 'Размеры товаров в наличии'
    ws['M1'] = 'Рейтинг'
    ws['N1'] = 'Количество отзывов'
    for i, info in enumerate(filtered_info, start=2):
        ws[f'A{i}'] = info['url']
        ws[f'B{i}'] = info['article']
        ws[f'C{i}'] = info['name']
        ws[f'D{i}'] = info['price']
        ws[f'E{i}'] = ", ".join(info['images'])
        ws[f'F{i}'] = info['description']
        ws[f'G{i}'] = ", ".join([f"{d['key']} - {d['value']}" for d in info['characteristic']["Основная информация"]])
        ws[f'H{i}'] = ", ".join([f"{d['key']} - {d['value']}" for d in info['characteristic']['Дополнительная информация']])
        ws[f'I{i}'] = info['seller_name']
        ws[f'J{i}'] = info['seller_link']
        ws[f'K{i}'] = ", ".join(info['sizes'])
        ws[f'L{i}'] = ", ".join(info['sizes_available'])
        ws[f'M{i}'] = info['rating']
        ws[f'N{i}'] = info['reviews_amount']

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()



